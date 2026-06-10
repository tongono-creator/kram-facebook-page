# -*- coding: utf-8 -*-
"""review.py — generate รูปรีวิวสินค้าจาก review_products.xlsx แล้วโพส FB"""

import sys, io, os, base64, requests, time, random, re
from datetime import datetime, timezone, timedelta
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from google import genai
import openpyxl
from overlay_utils import add_overlay

GEMINI_API_KEY    = os.environ.get("GEMINI_API_KEY",    "")
PAGE_ACCESS_TOKEN = os.environ.get("KRAM_PAGE_ACCESS_TOKEN", "")
PAGE_ID           = "116701184708556"
TEXT_MODELS       = ["gemini-2.5-flash", "gemini-3.5-flash"]
OUTPUT_DIR        = "output"
EXCEL_PATH        = os.path.join(os.path.dirname(__file__), "review_products.xlsx")
AFFILIATE_DIR     = os.path.join(os.path.dirname(__file__), "affiliate_data")
ACCENT_COLOR      = (0, 191, 255) # ฟ้า #00BFFF สำหรับกรามค้าง

if not GEMINI_API_KEY:
    try:
        from config import GEMINI_API_KEY, PAGE_ACCESS_TOKEN
    except ImportError:
        pass

os.makedirs(OUTPUT_DIR, exist_ok=True)
API_ENABLED = True
client = None
if not GEMINI_API_KEY or GEMINI_API_KEY in ("DUMMY_KEY", "DUMMY"):
    print("[Warning] GEMINI_API_KEY is not set or is a dummy key. Disabling API calls.")
    API_ENABLED = False
else:
    try:
        client = genai.Client(api_key=GEMINI_API_KEY, http_options={'timeout': 60000.0})
    except Exception as e:
        print(f"[Warning] Failed to initialize genai.Client: {e}. Disabling API calls.")
        API_ENABLED = False

def load_next_product():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=False):
        no    = row[0].value
        detail= row[1].value
        shopee= row[2].value
        lazada= row[3].value
        imgurl= row[4].value
        promo = row[5].value
        posted= row[6].value

        # ข้าม sample row และ posted แล้ว
        if not detail or "วางรายละเอียด" in str(detail):
            continue
        if str(posted).strip().lower() == "done" or str(posted).strip().startswith("done"):
            continue
        if not shopee or "xxx" in str(shopee):
            continue

        return {
            "row": row[0].row,
            "no": no,
            "detail": str(detail).strip(),
            "shopee": str(shopee).strip(),
            "lazada": str(lazada).strip() if lazada else "",
            "image_url": str(imgurl).strip() if imgurl else "",
            "promo": str(promo).strip() if promo else "",
        }, wb, ws
    return None, wb, ws

def mark_posted(wb, ws, row_num):
    bkk = timezone(timedelta(hours=7))
    ts = datetime.now(bkk).strftime("%Y-%m-%d %H:%M")
    ws.cell(row=row_num, column=7, value=f"done {ts}")
    wb.save(EXCEL_PATH)
    print(f"Marked row {row_num} as done")

def clean_promo(raw):
    """เอาเฉพาะบรรทัดที่มี ฿ หรือ ลด หรือ % หรือ ส่งฟรี"""
    if not raw:
        return ""
    lines = raw.strip().splitlines()
    kept = [l.strip() for l in lines if re.search(r'฿|ลด|%|ส่งฟรี|flash|sale', l, re.IGNORECASE)]
    return " | ".join(kept[:3]) if kept else ""

def get_allowed_xlsx_files():
    path = os.path.abspath(__file__).replace("\\", "/")
    if "chowchow" in path:
        return ["สัตว์เลี้ยง.xlsx"]
    elif "somtam" in path:
        return ["อาหารและเครื่องดื่ม.xlsx"]
    elif "rocket" in path:
        return ["เครื่องใช้ไฟฟ้าภายในบ้าน.xlsx"]
    elif "x-bot" in path:
        return ["สินค้าขายดี.xlsx"]
    else:  # kram-facebook-page
        return ["เครื่องใช้ในบ้าน.xlsx", "ค่าคอมพิเศษ.xlsx", "เสื้อผ้าแฟชั่นผู้หญิง.xlsx", "สินค้าสำหรับเม้นใต้คลิป.xlsx"]

def get_posted_urls(ws):
    urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 4:
            continue
        shopee = row[2]
        lazada = row[3]
        if shopee and str(shopee).strip().startswith("http"):
            urls.add(str(shopee).strip())
        if lazada and str(lazada).strip().startswith("http"):
            urls.add(str(lazada).strip())
    return urls

def append_posted_fallback(wb, ws, product):
    bkk = timezone(timedelta(hours=7))
    ts = datetime.now(bkk).strftime("%Y-%m-%d %H:%M")
    row_num = ws.max_row + 1
    
    ws.cell(row=row_num, column=1, value=product["no"])
    ws.cell(row=row_num, column=2, value=product["detail"])
    ws.cell(row=row_num, column=3, value=product["shopee"])
    ws.cell(row=row_num, column=4, value=product["lazada"])
    ws.cell(row=row_num, column=5, value=product["image_url"])
    ws.cell(row=row_num, column=6, value=product["promo"])
    ws.cell(row=row_num, column=7, value=f"done {ts}")
    
    wb.save(EXCEL_PATH)
    print(f"Appended fallback product to review_products.xlsx at row {row_num}")

def load_affiliate_product(posted_urls):
    """Fallback: สุ่มสินค้าจาก AFFILIATE_DIR เมื่อ review_products.xlsx หมด"""
    import glob
    allowed_names = get_allowed_xlsx_files()
    xlsx_files = []
    for name in allowed_names:
        p = os.path.join(AFFILIATE_DIR, name)
        if os.path.exists(p):
            xlsx_files.append(p)
            
    # Fallback to any xlsx files if no mapped files exist
    if not xlsx_files:
        xlsx_files = glob.glob(os.path.join(AFFILIATE_DIR, "*.xlsx"))
        
    if not xlsx_files:
        print(f"[affiliate] No xlsx files found in {AFFILIATE_DIR}")
        return None
        
    random.shuffle(xlsx_files)
    for xlsx_path in xlsx_files:
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            candidates = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 10:
                    continue
                name   = row[1]
                imgurl = row[2]
                price  = row[3]
                shopee = row[9]
                lazada = row[10] if len(row) > 10 else None
                if not shopee or not name:
                    continue
                
                shopee_val = str(shopee).strip()
                lazada_val = str(lazada).strip() if lazada else ""
                
                # Check against posted URLs
                if shopee_val in posted_urls or (lazada_val and lazada_val in posted_urls):
                    continue
                    
                candidates.append({
                    "no": row[0],
                    "detail": f"{name} ราคา {price} บาท",
                    "shopee": shopee_val,
                    "lazada": lazada_val,
                    "image_url": str(imgurl).strip() if imgurl else "",
                    "promo": "",
                    "row": None,
                })
            wb.close()
            if candidates:
                product = random.choice(candidates)
                print(f"[affiliate] Loaded from {os.path.basename(xlsx_path)}: {product['detail'][:60]}")
                return product
        except Exception as e:
            print(f"[affiliate] Failed to read {xlsx_path}: {e}")
    print("[affiliate] No valid product found in any xlsx file")
    return None


def extract_highlights(detail, promo):
    """ให้ AI สกัดจุดเด่นจาก raw detail"""
    highlights = None
    prompt = (
        f"จากรายละเอียดสินค้านี้:\n{detail}\n\n"
        f"สกัดจุดเด่นสินค้าเป็นประโยคข้อความสั้นแนวธรรมชาติ 1-2 ย่อหน้าสั้นๆ (ห้ามทำเป็นข้อๆ หรือมีสัญลักษณ์รายการ/bullet points เช่น •, ▪️, - หรือเลขข้อ) "
        f"เน้นประโยชน์ที่คนซื้อสนใจและใช้งานจริง ห้ามใส่ข้อมูลราคาหรือโปรโมชั่น "
        f"ตอบเฉพาะส่วนรายละเอียดเนื้อความเท่านั้น"
    )
    for model in TEXT_MODELS:
        try:
            resp = client.models.generate_content(model=model, contents=prompt)
            highlights = resp.text.strip()
            if highlights:
                break
        except Exception as e:
            err_msg = str(e)
            print(f"[{model}] highlights failed: {err_msg[:80]}")
            if any(x in err_msg.lower() for x in ["api key not valid", "permission_denied", "api_key_invalid"]):
                break
    if not highlights:
        print("[Warning] Highlights AI failed, using local fallback.")
        lines = [l.strip() for l in detail.splitlines() if l.strip()]
        points = []
        for line in lines:
            cleaned = re.sub(r'^[•\-\*\d\.\s–]+', '', line).strip()
            if cleaned and 5 < len(cleaned) < 100:
                points.append(cleaned)
            if len(points) >= 3:
                break
        if not points:
            points = [line[:80] for line in lines[:2]]
        highlights = " ".join(points) if points else "รายละเอียดเพิ่มเติมศึกษาต่อได้ที่หน้าร้านเลยครับ"
    if promo:
        highlights += f"\n🔥 โปรตอนนี้: {promo}"
    return highlights

def download_image(url):
    """Download รูปแรกจาก URL"""
    first_url = url.strip().split("\n")[0].strip()
    resp = requests.get(first_url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
    if resp.status_code == 200:
        bkk = timezone(timedelta(hours=7))
        ts = datetime.now(bkk).strftime("%Y%m%d_%H%M%S")
        ext = "webp" if "webp" in first_url else "jpg"
        path = os.path.join(OUTPUT_DIR, f"product_{ts}.{ext}")
        with open(path, "wb") as f:
            f.write(resp.content)
        print(f"Product image downloaded: {path}")
        return path
    raise RuntimeError(f"Image download failed: {resp.status_code}")

def generate_hook(detail, highlights):
    """สร้างหัวข้อสั้นพาดหัวรูปภาพ 2 บรรทัด คั่นด้วย '|'"""
    prompt = (
        f"จากรายละเอียดสินค้าต่อไปนี้:\n{detail}\n\n"
        f"จุดเด่นสินค้าที่สกัดแล้ว:\n{highlights}\n\n"
        "กรุณาสร้างคำพาดหัวโฆษณารีวิวสินค้านี้ภาษาไทยสั้นๆ 2 บรรทัด คั่นด้วยเครื่องหมาย pipe '|' (บรรทัด 1 | บรรทัด 2)\n"
        "กฎในการร่าง:\n"
        "- บรรทัด 1: คำโปรย/ชื่อเล่นสุดปังสไตล์วัยรุ่นหรือคนทำงานขำขัน (3-5 คำ)\n"
        "- บรรทัด 2: เหตุผลโดนใจ/จุดเด่นในการแก้ปัญหา (4-7 คำ)\n"
        "- ห้ามใช้เครื่องหมายคำพูด อัญประกาศ หรือข้อความนำหน้า/ตามหลังใดๆ\n"
        "- ห้ามมี Emoji ปนในหัวข้อนี้เด็ดขาด\n"
        "ตัวอย่าง: เบาะรองนั่งสู้ชีวิต | นั่งทำงาน 10 ชม. ไม่ปวดหลัง"
    )
    for model in TEXT_MODELS:
        try:
            resp = client.models.generate_content(model=model, contents=prompt)
            result = resp.text.strip()
            label_pattern = r'^(ข้อความในโพสต์\s*Facebook|Facebook\s*Caption|Facebook\s*caption|Caption|caption|ข้อความบนรูป|ข้อความในรูป|ข้อความ|คำบรรยาย|คำอธิบาย|บรรทัดที่\s*\d+|บรรทัด\s*\d+|ประโยคที่\s*\d+|ประโยค\s*\d+|Hook\s*text|Hook|Line\s*\d+|[L|l]ine\s*\d+|\d+)\s*[:\-\.\s]\s*'
            result = re.sub(label_pattern, '', result, flags=re.IGNORECASE).strip()
            result = result.strip('"\'“”‘’')
            if "|" in result:
                parts = result.split("|", 1)
                line1 = parts[0].strip()
                line2 = parts[1].strip()
                return line1, line2
            else:
                return result[:15], ""
        except Exception as e:
            print(f"[{model}] hook generation failed: {e}")
    # Smart local fallback for hook
    first_line = detail.split('\n')[0].strip()
    first_line = re.sub(r'[\[\]\(\)]', '', first_line)
    first_line = re.sub(r'^[•\-\*\d\.\s–]+', '', first_line).strip()
    _specs = ['ราคา', 'แพ็ค', 'ขนาด', 'สำหรับ', 'จำนวน', 'กรัม', 'ลิตร', ' ml', ' kg', ' g ', ',']
    _stop = len(first_line)
    for _kw in _specs:
        _idx = first_line.find(_kw)
        if _idx > 2:
            _stop = min(_stop, _idx)
    short_title = first_line[:_stop].strip()
    words = short_title.split()
    line1 = "สินค้าแนะนำ"
    for n in (3, 2, 1):
        candidate = " ".join(words[:n]) if words else "สินค้าแนะนำ"
        if len(candidate) <= 20 or n == 1:
            line1 = candidate
            break
    price_m = re.search(r'ราคา\s*([\d,]+(?:\.\d+)?)\s*บาท', detail)
    if price_m:
        line2 = f"ราคา {price_m.group(1)} บาท"
    else:
        feat_lines = [l.strip() for l in detail.splitlines() if l.strip() and l.strip() != first_line.strip()]
        feat_lines = [re.sub(r'^[•\-\*\d\.\s–]+', '', l).strip() for l in feat_lines]
        feat_lines = [l for l in feat_lines if 5 < len(l) < 60]
        line2 = feat_lines[0][:20] if feat_lines else ""
    return line1, line2

def generate_caption(detail, shopee, lazada, promo, highlights):
    global API_ENABLED
    import random
    import hashlib as _hs
    import re
    
    is_x = "x-bot" in __file__.replace("\\", "/")
    promo_line = f"\n🔥 โปรโมชั่น: {promo}" if promo else ""
    if is_x:
        promo_line = f" 🔥 {promo}" if promo else ""
        
    caption = None
    
    # Archetypes with weights: 30% Friend, 30% Review, 20% Promo, 20% Life Story
    archetypes = [
        ("เพื่อนที่มาบอกต่อ แนะนำของดีให้เพื่อน", "เพื่อนบอกต่อ"),
        ("คนที่ชอบซื้อของออนไลน์และมารีวิวสั้นๆ หลังใช้งานจริงมาระยะหนึ่ง", "รีวิวหลังใช้"),
        ("คนที่บังเอิญเจอราคาโปรโมชั่นหรือส่วนลดพิเศษแล้วอยากเอามาแชร์ต่อ", "เจอโปรมาแชร์"),
        ("คนที่บ่นหรือเล่าเรื่องราวชีวิตประจำวัน/อุปสรรคชีวิตทั่วไปก่อน แล้วโยงเข้าหาตัวสินค้าที่เข้ามาช่วยแก้ปัญหา", "เล่าเรื่องชีวิตแล้วโยงเข้าสินค้า")
    ]
    weights = [30, 30, 20, 20]
    
    selected_arch_desc, selected_arch_name = random.choices(archetypes, weights=weights, k=1)[0]
    
    path_norm = __file__.replace("\\", "/").lower()
    if "somtam" in path_norm:
        gender_inst = f"คุณคือคนธรรมดาที่ซื้อของออนไลน์บ่อย และชอบแชร์ของที่คิดว่าคุ้มให้เพื่อน โดยรอบนี้สุ่มบทบาทเป็น: {selected_arch_desc} ใช้คำลงท้ายว่า 'ค่ะ' หรือ 'คะ' และสรรพนามแทนตัวว่า 'หนู' หรือ 'เรา' เท่านั้น"
        closing = "ดูลิ้งในคอมเมนต์แรกเลยนะคะ 👇"
        closing_fallback_promo = "แปะพิกัดไว้ในคอมเมนต์แรกเลยนะคะ 👇"
        fallbacks = [
            "เจอ {title_clean} อันนี้มาลองใช้แล้วดีงามมาก แนะนำค่า{_price_str}\n\nดูลิ้งในคอมเมนต์แรกเลยนะคะ 👇",
            "ใครหา {title_clean} อยู่ ตัวนี้ลองใช้มาระยะนึงแล้ว ดีกว่าที่คิดไว้เยอะเลย{_price_str}\n\nดูลิ้งในคอมเมนต์แรกเลยนะคะ 👇",
            "บังเอิญเจอ {title_clean} ลดเหลือแค่นี้เอง แปะพิกัดไว้เผื่อใครตามหาอยู่เนอะ{_price_str}\n\nดูลิ้งในคอมเมนต์แรกเลยนะคะ 👇",
            "ช่วงนี้เจอปัญหาชีวิตประจำวันนิดหน่อย ดีที่ได้ {title_clean} ตัวนี้มาช่วย สะดวกขึ้นเยอะเลยค่ะ{_price_str}\n\nดูลิ้งในคอมเมนต์แรกเลยนะคะ 👇"
        ]
    elif "chowchow" in path_norm:
        gender_inst = f"คุณคือคนธรรมดาที่ซื้อของออนไลน์บ่อย และชอบแชร์ของที่คิดว่าคุ้มให้เพื่อน โดยรอบนี้สุ่มบทบาทเป็น: {selected_arch_desc} ใช้คำลงท้ายว่า 'ฮะ' หรือ 'โฮ่ง' และสรรพนามแทนตัวว่า 'น้องตูบ' หรือ 'ผม' เท่านั้น และมีกลิ่นอายความซนแบบน้องหมา"
        closing = "กดลิ้งในคอมเมนต์แรกได้เลยฮะ 👇"
        closing_fallback_promo = "พิกัดอยู่ในคอมเมนต์แรกนะโฮ่ง 👇"
        fallbacks = [
            "เจอ {title_clean} อันนี้มาลองใช้แล้วชอบมาก แนะนำฮะ{_price_str}\n\nกดลิ้งในคอมเมนต์แรกได้เลยฮะ 👇",
            "ใครหา {title_clean} อยู่ ตัวนี้ลองใช้มาระยะนึงแล้ว ดีกว่าที่คิดไว้เยอะเลย{_price_str}\n\nกดลิ้งในคอมเมนต์แรกได้เลยฮะ 👇",
            "บังเอิญเจอ {title_clean} ลดราคาเหลือเท่านี้ แปะพิกัดไว้เผื่อใครหาอยู่ฮะ{_price_str}\n\nกดลิ้งในคอมเมนต์แรกได้เลยฮะ 👇",
            "ช่วงนี้เจอปัญหาวุ่นๆ ดีที่ได้ {title_clean} ตัวนี้มาช่วย ชีวิตง่ายขึ้นเยอะโฮ่ง{_price_str}\n\nกดลิ้งในคอมเมนต์แรกได้เลยฮะ 👇"
        ]
    elif "x-bot" in path_norm:
        gender_inst = f"คุณคือคนธรรมดาที่ซื้อของออนไลน์บ่อย และมาโพสต์สั้นๆ บน X (Twitter) เล่าแบบเพื่อนคุยกัน โดยรอบนี้สุ่มบทบาทเป็น: {selected_arch_desc} ใช้คำลงท้าย 'ครับ' สรรพนาม 'ผม'"
        closing = ""
        closing_fallback_promo = ""
        fallbacks = [
            "เจอ {title_clean} มาลองใช้แล้วชอบเลย แนะนำครับ{_price_str}",
            "ใครหา {title_clean} อยู่ ลองใช้มาระยะนึงแล้ว รู้สึกดีกว่าที่คิดไว้ครับ{_price_str}",
            "บังเอิญเจอ {title_clean} ลดราคาเหลือเท่านี้ แปะพิกัดไว้เผื่อใครหาอยู่{_price_str}",
            "ชีวิตช่วงนี้สบายขึ้นเยอะเพราะได้ {title_clean} ตัวนี้มาช่วย สะดวกมากครับ{_price_str}"
        ]
    else:  # rocket & kram
        gender_inst = f"คุณคือคนธรรมดาที่ซื้อของออนไลน์บ่อย และชอบแชร์ของที่คิดว่าคุ้มให้เพื่อน โดยรอบนี้สุ่มบทบาทเป็น: {selected_arch_desc} ใช้คำลงท้ายว่า 'ครับ' และสรรพนามแทนตัวว่า 'ผม' หรือ 'พี่' เท่านั้น"
        closing = "ดูลิ้งในคอมเมนต์แรกเลยครับ 👇"
        closing_fallback_promo = "แปะพิกัดไว้ในคอมเมนต์แรกเลยครับ 👇"
        fallbacks = [
            "เจอ {title_clean} อันนี้มาลองใช้แล้วโอเคเลย แนะนำครับ{_price_str}\n\nดูลิ้งในคอมเมนต์แรกเลยครับ 👇",
            "ใครหา {title_clean} อยู่ ตัวนี้ลองใช้มาระยะนึงแล้ว รู้สึกดีกว่าที่คิดไว้ครับ{_price_str}\n\nดูลิ้งในคอมเมนต์แรกเลยครับ 👇",
            "บังเอิญเจอ {title_clean} ลดราคาเหลือเท่านี้ แปะพิกัดไว้เผื่อใครหาอยู่ครับ{_price_str}\n\nดูในคอมเมนต์แรกเลยครับ 👇",
            "ช่วงนี้พยายามแก้ปัญหาชีวิตประจำวัน พอดีได้ {title_clean} ตัวนี้มา สะดวกขึ้นเยอะครับ{_price_str}\n\nดูในคอมเมนต์แรกเลยครับ 👇"
        ]

    title_lines = [l.strip() for l in detail.splitlines() if l.strip()]
    title_raw = title_lines[0] if title_lines else ""
    title_clean = re.sub(r'^[•\-\*\d\.\s\u2013\(\[\{\)\|\}]+', '', title_raw).strip()
    title_clean = title_clean[:50]
    
    price_m = re.search(r'ราคา\s*([\d,]+(?:\.\d+)?)\s*บาท', detail)
    price_val = price_m.group(1) if price_m else ""
    _price_str = f" ราคาแค่ {price_val} บาท" if price_val else ""
    if is_x:
        _price_str = f" ราคา {price_val} บาท" if price_val else ""

    active_client = globals().get("client")
    if API_ENABLED and active_client:
        prompt = (
            f"{gender_inst}\n\n"
            "ก่อนเขียนโพสต์ ให้คิดก่อนว่า:\n"
            "- สินค้านี้แก้ปัญหาอะไร\n"
            "- คนซื้อเพราะอะไร\n"
            "- จุดไหนที่ทำให้รู้สึกว่า 'เออ น่าสนใจ'\n\n"
            f"รายละเอียดสินค้า:\n{detail}\n\n"
            "กฎเหล็กสำคัญมาก:\n"
            "- ห้ามเปิดโพสต์ด้วยชื่อสินค้า หรือแบรนด์เด็ดขาด\n"
            "- ห้ามคัดลอกชื่อสินค้าเต็มจาก Shopee\n"
            "- ห้ามใส่สเปกยาวๆ\n"
            "- ห้ามใช้คำขายของเช่น 'คุ้มมาก', 'ของมันต้องมี', 'รีบซื้อ', 'โปรโมชั่น', 'สั่งได้เลย', 'อย่าพลาด', 'ไม่ควรพลาด'\n"
            "- ให้เขียนเหมือนคนใช้จริงมาเล่า เป็นกันเองและเป็นธรรมชาติที่สุด\n\n"
        )
        
        if is_x:
            prompt += (
                "รูปแบบโพสต์บน X (Twitter):\n"
                "1. เปิดด้วยปัญหาหรือความรู้สึกสั้นๆ\n"
                "2. พูดถึงสินค้าสั้นๆ (ใช้ชื่อย่อ/แบรนด์ย่อ)\n"
                "3. บอกจุดเด่นสั้นๆ 1 ข้อ\n\n"
                "ความยาว: 2-3 ประโยค ห้ามเกิน 150 ตัวอักษรรวม\n"
                "ตอบเฉพาะตัวโพสต์เท่านั้น"
            )
        else:
            prompt += (
                "รูปแบบโพสต์บน Facebook:\n"
                "1. เปิดด้วยปัญหาหรือความรู้สึก\n"
                "2. เล่าว่าเจออะไร\n"
                "3. พูดถึงสินค้าสั้นๆ (ใช้ชื่อแบรนด์หรือชื่อย่อเท่านั้น ห้ามใช้ชื่อเต็ม)\n"
                "4. บอกเหตุผลที่ชอบ 1-2 จุด จากมุมผู้ใช้จริง\n"
                f"5. ปิดท้ายกระตุ้นการกระทำด้วยประโยคว่า: '{closing}'\n\n"
                "ความยาว: 2-5 ประโยค\n"
                "ตอบเฉพาะตัวโพสต์เท่านั้น ไม่ต้องพูดนำหน้า/อธิบายใดๆ"
            )
            
        for model in TEXT_MODELS:
            try:
                resp = active_client.models.generate_content(model=model, contents=prompt)
                caption_text = resp.text.strip()
                if caption_text:
                    lines = caption_text.splitlines()
                    while lines and (
                        re.search(r'^(ได้เลย|นี่คือ|แน่นอน|โพสต์รีวิว|ครับ|ค่ะ|---)', lines[0].strip(), re.IGNORECASE)
                        or lines[0].strip() in ("", "---")
                    ):
                        lines.pop(0)
                    caption = "\n".join(lines).strip()
                    break
            except Exception as e:
                err_msg = str(e)
                print(f"[{model}] caption generation failed: {err_msg[:80]}")
                if any(x in err_msg.lower() for x in ["api key", "invalid_argument", "api_key", "timeout", "timed out", "deadline exceeded", "connection", "connect", "unreachable"]):
                    print("Persistent API key or network/timeout issue detected. Disabling API calls immediately.")
                    API_ENABLED = False
                    break
                    
        if not caption and API_ENABLED:
            print("[Warning] Caption generation failed on all models. Disabling API calls for this run.")
            API_ENABLED = False

    if not caption:
        print("[Warning] Falling back to local heuristic caption.")
        arch_idx = 0
        if selected_arch_name == "เพื่อนบอกต่อ":
            arch_idx = 0
        elif selected_arch_name == "รีวิวหลังใช้":
            arch_idx = 1
        elif selected_arch_name == "เจอโปรมาแชร์":
            arch_idx = 2
        else:
            arch_idx = 3
            
        template = fallbacks[arch_idx]
        caption = template.format(title_clean=title_clean or "ของดี", _price_str=_price_str)
        
    if promo:
        caption += promo_line
        
    return caption

def _post_one_comment(post_id, text):
    try:
        resp = requests.post(
            f"https://graph.facebook.com/v21.0/{post_id}/comments",
            data={"access_token": PAGE_ACCESS_TOKEN, "message": text},
            timeout=30
        )
        result = resp.json()
        if "id" in result:
            print(f"Comment posted: {result['id']}")
        else:
            print(f"Comment failed: {result}")
    except Exception as e:
        print(f"Comment error: {e}")

def post_link_comment(post_id, shopee, lazada, promo):
    """โพส comment ลิ้งใต้โพส แยก Shopee / Lazada คนละคอมเม้น"""
    promo_line = f"\n🔥 โปร: {promo}" if promo else ""
    if shopee and "xxx" not in shopee:
        _post_one_comment(post_id, f"👉 ซื้อได้ที่ Shopee → {shopee}{promo_line}")
    if lazada and "xxx" not in lazada:
        _post_one_comment(post_id, f"🛍️ หรือสั่งทาง Lazada → {lazada}")

def post_to_page(img_path, caption, shopee=None, lazada=None, promo=None, scheduled_timestamp=None):
    print("Posting to Facebook Page...")
    from affiliate_utils import get_next_scheduled_time

    if scheduled_timestamp is not None:
        scheduled_time = scheduled_timestamp
    else:
        slots = ["08:00"]
        scheduled_time = get_next_scheduled_time(slots)
    
    if scheduled_time:
        comment_texts = []
        promo_line = f"\n🔥 โปร: {promo}" if promo else ""
        if shopee and "xxx" not in shopee:
            comment_texts.append(f"👉 ซื้อได้ที่ Shopee → {shopee}{promo_line}")
        if lazada and "xxx" not in lazada:
            comment_texts.append(f"🛍️ หรือสั่งทาง Lazada → {lazada}")
            
        if comment_texts:
            caption += "\n\n📌 ชี้เป้าของดีน่าสนใจ:\n" + "\n".join(comment_texts)
            
        print(f"Scheduling to Facebook for timestamp {scheduled_time}...")
        with open(img_path, "rb") as f:
            resp = requests.post(
                f"https://graph.facebook.com/v25.0/{PAGE_ID}/photos",
                data={
                    "access_token": PAGE_ACCESS_TOKEN,
                    "message": caption,
                    "published": "false",
                    "unpublished_content_type": "SCHEDULED",
                    "scheduled_publish_time": scheduled_time
                },
                files={"source": ("review.png", f, "image/png")},
                timeout=60
            )
        result = resp.json()
        if "id" in result:
            photo_id = result.get("post_id") or result["id"]
            print(f"Scheduled successfully! Photo ID: {photo_id}")
            return photo_id, True
        else:
            print(f"FB Error: {result}")
            raise SystemExit(1)

    with open(img_path, "rb") as f:
        resp = requests.post(
            f"https://graph.facebook.com/v25.0/{PAGE_ID}/photos",
            data={"access_token": PAGE_ACCESS_TOKEN, "message": caption, "published": "true"},
            files={"source": ("review.png", f, "image/png")},
            timeout=60
        )
    result = resp.json()
    if "id" in result:
        post_id = result.get("post_id") or result["id"]
        print(f"Page Posted! ID: {post_id}")
        print(f"https://www.facebook.com/{post_id}")
        return post_id, False
    else:
        print(f"FB Error: {result}")
def extract_badge_text(promo):
    if not promo:
        return None
    pct_match = re.search(r'(ลด\s*\d+\s*%)|(\d+\s*%\s*OFF)', promo, re.IGNORECASE)
    if pct_match:
        val = pct_match.group(0)
        val = re.sub(r'\s+', ' ', val)
        return val
    price_match = re.search(r'฿\s*\d+', promo)
    if price_match:
        return price_match.group(0).replace(" ", "")
    return None


if __name__ == "__main__":
    import argparse
    import time as _time

    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Run without posting or marking as done")
    args = parser.parse_args()

    IMMEDIATE = os.environ.get("IMMEDIATE", "false").lower() == "true"

    bkk = timezone(timedelta(hours=7))
    now_bkk = datetime.now(bkk)

    # Pre-calculate 5 timestamps: 05:00/08:00/11:00/14:00/17:00 BKK
    # IMMEDIATE=true (workflow_dispatch) -> post now, no scheduling
    slot_times = ["05:00", "08:00", "11:00", "14:00", "17:00"]
    def make_slot_ts(slot_str):
        if IMMEDIATE:
            return None
        h, m = map(int, slot_str.split(":"))
        dt = now_bkk.replace(hour=h, minute=m, second=0, microsecond=0)
        if dt <= now_bkk:
            dt += timedelta(days=1)
        return int(dt.timestamp())

    slot_timestamps = [make_slot_ts(s) for s in slot_times]
    posted_this_run = set()

    for i, sched_ts in enumerate(slot_timestamps):
        print(f"\n--- Post {i+1}/5 (slot {slot_times[i]}) ---")

        product, wb, ws = load_next_product()
        affiliate_mode = False
        if not product:
            print("review_products.xlsx หมดแล้ว — ลอง fallback จาก AFFILIATE_DIR")
            posted_urls = get_posted_urls(ws)
            product = load_affiliate_product(posted_urls)
            affiliate_mode = True
            if not product:
                print("ไม่มีสินค้าเหลือ หยุด")
                break

        if product.get("shopee") in posted_this_run:
            print(f"[Skip] ซ้ำใน run นี้: {str(product.get('shopee',''))[:60]}")
            continue

        print(f"Product: {product['detail'][:60]}...")

        promo_clean = clean_promo(product["promo"])
        highlights  = extract_highlights(product["detail"], promo_clean)

        line1, line2 = generate_hook(product["detail"], highlights)
        print(f"Hook: {line1} | {line2}")

        product_img = download_image(product["image_url"])

        try:
            badge_text = extract_badge_text(product.get("promo"))
            review_img = add_overlay(
                product_img, line1, line2, ACCENT_COLOR,
                font_name="Itim-Regular.ttf",
                badge_text=badge_text,
                watermark="คราม Kram"
            )
            os.unlink(product_img)
            print(f"Review image overlaid: {review_img}")
        except Exception as overlay_err:
            print(f"Overlay failed, using original image: {overlay_err}")
            review_img = product_img

        caption = generate_caption(
            product["detail"], product["shopee"],
            product["lazada"], promo_clean, highlights
        )
        print(f"Caption:\n{caption}\n")

        if args.dry_run:
            print(f"Dry run post {i+1}. image={review_img}")
            print(f"Link: 👉 Shopee → {product['shopee']}")
            posted_this_run.add(product.get("shopee"))
        else:
            post_id, was_scheduled = post_to_page(
                review_img, caption,
                product["shopee"], product["lazada"], promo_clean,
                scheduled_timestamp=sched_ts
            )
            posted_this_run.add(product.get("shopee"))
            if os.path.exists(review_img):
                os.unlink(review_img)
            if not was_scheduled:
                post_link_comment(post_id, product["shopee"], product["lazada"], promo_clean)
            if not affiliate_mode:
                mark_posted(wb, ws, product["row"])
            else:
                append_posted_fallback(wb, ws, product)
            _time.sleep(5)

